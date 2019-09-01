
from openpyxl import load_workbook
wb2 = load_workbook('/home/f2m4/mycode/qhj/devapi/ziliao/生产保障中心人员职工编号.xlsx')
sheet_ranges = wb2['中秋节']
for row in sheet_ranges.rows:
    for cell in row:
        print(cell.value)


# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# ws.append([1, 2, 3])
# # Save the file
# wb.save("sample.xlsx")