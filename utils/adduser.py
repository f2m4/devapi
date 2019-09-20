from django.contrib.auth.models import User
from openpyxl import load_workbook

def godb():
    wb2 = load_workbook('ziliao/生产保障中心人员职工编号.xlsx')
    sheet_ranges = wb2['中秋节']
    rows_iter= (row for row in sheet_ranges.rows)



    def db_create_user(username,first_name, password, is_active):
        if is_active == '0':
            is_active = False
        elif is_active == '1':
            is_active = True
        try:
            User.objects.create_user(username=username, password=password, is_active=is_active,first_name=first_name,)
        except Exception as e:
            return e   # 出现错误
        else:
            return '完成'    # 创建成功

    # def add_list_user(username,first_name, password, is_active):
    #     createResult = db_create_user(username, first_name,password, is_active)
    #     if createResult == 1:
    #         return "{}->创建成功".format(username)
    #     elif createResult == -1:
    #         return "{}->创建失败".format(username)
    i=1
    for row in rows_iter:
        cell_list=[i.value for i in row]
        print('开始导入第 {} 行'.format(i,cell_list[0]))
        print(cell_list)
        if cell_list[0] is None:
            print('!!!!!!!!!!!!!!!!!!!跳过此行!!!!!!!!!')
            i+=1
            continue
        i+=1
        print(db_create_user(cell_list[1],cell_list[0],"88888888",'1'))
        print('导入完毕~~~~~~~~~~~~~~~~~~')


# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# ws.append([1, 2, 3])
# # Save the file
# wb.save("sample.xlsx")