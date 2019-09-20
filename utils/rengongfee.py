from openpyxl import load_workbook
from internalmarket import models

def godb():
    wb2 = load_workbook('ziliao/定额表.xlsx')
    # sheet_ranges = wb2['1']
    # sheet_ranges = wb2['2']
    sheet_ranges = wb2['3']
    rows_iter= (row for row in sheet_ranges.rows)

    def db_create_go(name,workingHours, fee,remark):
        try:
            obj=models.QingGongModel.objects.create(name=name, workingHours=workingHours, fee=fee,remark=remark)
        except Exception as e:
            print("{}->失败原因".format(e))
            return "{}->创建失败".format(name)
        else:
            tag = models.TagModel.objects.get_or_create(name=remark)[0]
            obj.tag.add(tag)
            return "{}->创建成功".format(name)   # 创建成功

    i = 1
    for row in rows_iter:
        cell_list = [i.value for i in row]
        print('开始导入第 {} 行'.format(i, cell_list[0]))
        # print(cell_list)
        if cell_list[0] is None:
            print('!!!!!!!!!!!!!!!!!!!跳过此行!!!!!!!!!')
            i += 1
            continue
        i += 1
        print(db_create_go(cell_list[1], cell_list[4] if cell_list[4] is not None else '', cell_list[5],cell_list[6]))
        print('导入完毕~~~~~~~~~~~~~~~~~~')


    print('添加完成')
# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# ws.append([1, 2, 3])
# # Save the file
# wb.save("sample.xlsx")